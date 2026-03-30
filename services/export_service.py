"""
Export Service - Handle PDF and Excel exports for reports and data
"""
import os
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Service for exporting data to PDF and Excel formats"""
    
    @staticmethod
    def export_student_report_pdf(student_data):
        """
        Generate PDF report for a single student
        
        Args:
            student_data: Dict containing student info, marks, attendance, etc.
        
        Returns:
            BytesIO: PDF file object
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066cc'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#003366'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph(f"Student Performance Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Student Info
        elements.append(Paragraph("Student Information", heading_style))
        student_info = [
            ['Field', 'Value'],
            ['Name', student_data.get('name', 'N/A')],
            ['Roll Number', student_data.get('roll_number', 'N/A')],
            ['Email', student_data.get('email', 'N/A')],
            ['Department', student_data.get('department', 'N/A')],
            ['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        student_table = Table(student_info, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Academic Performance
        if 'marks' in student_data and student_data['marks']:
            elements.append(Paragraph("Academic Performance", heading_style))
            marks_data = [['Subject', 'Marks', 'Percentage', 'Grade']]
            
            for mark in student_data['marks']:
                marks_data.append([
                    mark.get('subject', 'N/A'),
                    str(mark.get('marks', 'N/A')),
                    str(mark.get('percentage', 'N/A')) + '%',
                    mark.get('grade', 'N/A')
                ])
            
            marks_table = Table(marks_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
            marks_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(marks_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Attendance
        if 'attendance' in student_data:
            elements.append(Paragraph("Attendance Summary", heading_style))
            attendance_data = [
                ['Metric', 'Value'],
                ['Total Classes', str(student_data['attendance'].get('total_classes', 0))],
                ['Classes Attended', str(student_data['attendance'].get('attended', 0))],
                ['Attendance %', str(student_data['attendance'].get('percentage', 0)) + '%'],
                ['Status', student_data['attendance'].get('status', 'N/A')]
            ]
            
            attendance_table = Table(attendance_data, colWidths=[2*inch, 4*inch])
            attendance_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(attendance_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Readiness Score
        if 'readiness' in student_data:
            elements.append(Paragraph("Placement Readiness", heading_style))
            readiness = student_data['readiness']
            readiness_data = [
                ['Metric', 'Score'],
                ['Readiness Score', str(readiness.get('score', 0)) + '/100'],
                ['Risk Level', readiness.get('risk_level', 'N/A')],
                ['Placement Outlook', readiness.get('placement_outlook', 'N/A')]
            ]
            
            readiness_table = Table(readiness_data, colWidths=[2*inch, 4*inch])
            readiness_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(readiness_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_class_report_pdf(class_data, class_name):
        """Generate PDF report for entire class"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#0066cc'),
            spaceAfter=20,
            alignment=1
        )
        
        # Title
        elements.append(Paragraph(f"Class Performance Report - {class_name}", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary Statistics
        if 'statistics' in class_data:
            stats = class_data['statistics']
            stats_data = [
                ['Metric', 'Value'],
                ['Total Students', str(stats.get('total_students', 0))],
                ['Average Marks', f"{stats.get('avg_marks', 0):.2f}"],
                ['Average Attendance', f"{stats.get('avg_attendance', 0):.2f}%"],
                ['At-Risk Students', str(stats.get('at_risk_count', 0))],
                ['Top Performers', str(stats.get('top_performers', 0))]
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Student List
        if 'students' in class_data and class_data['students']:
            elements.append(Paragraph("Student Performance Summary", styles['Heading2']))
            student_rows = [['Roll No', 'Name', 'Avg Marks', 'Attendance %', 'Status']]
            
            for student in class_data['students'][:30]:  # Limit to 30 per page
                student_rows.append([
                    student.get('roll_number', 'N/A')[:6],
                    student.get('name', 'N/A')[:20],
                    f"{student.get('avg_marks', 0):.1f}",
                    f"{student.get('attendance', 0):.1f}%",
                    student.get('status', 'N/A')
                ])
            
            student_table = Table(student_rows, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.3*inch, 1.5*inch])
            student_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(student_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_excel(data_list, headers, filename_prefix="export"):
        """
        Export data to Excel workbook
        
        Args:
            data_list: List of dicts or tuples containing data rows
            headers: List of column headers
            filename_prefix: Prefix for the filename
        
        Returns:
            BytesIO: Excel file object
        """
        workbook = Workbook()
        worksheet = workbook.active  # type: ignore
        if worksheet:
            worksheet.title = "Report"
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)  # type: ignore
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row_num, row_data in enumerate(data_list, 2):
            if isinstance(row_data, dict):
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)  # type: ignore
                    cell.value = row_data.get(header, '')
            else:
                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)  # type: ignore
                    cell.value = value
            
            # Alternate row colors
            if row_num % 2 == 0:
                for col_num in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)  # type: ignore
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):  # type: ignore
                try:
                    if len(str(row[0].value)) > max_length:
                        max_length = len(str(row[0].value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width  # type: ignore
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(data_list) + 1, min_col=1, max_col=len(headers)):  # type: ignore
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_students_excel(students, include_fields=None):
        """Export student list to Excel"""
        if include_fields is None:
            include_fields = ['name', 'roll_number', 'email', 'department', 'cgpa', 'attendance']
        
        buffer = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active  # type: ignore
        if worksheet:
            worksheet.title = "Students"
        
        # Headers
        headers = [field.replace('_', ' ').title() for field in include_fields]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)  # type: ignore
            cell.value = header  # type: ignore
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        # Data rows
        for row_num, student in enumerate(students, 2):
            for col_num, field in enumerate(include_fields, 1):
                cell = worksheet.cell(row=row_num, column=col_num)  # type: ignore
                cell.value = student.get(field, '') if isinstance(student, dict) else getattr(student, field, '')  # type: ignore
                
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Auto-adjust columns
        for col_num, header in enumerate(headers, 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = len(header) + 5  # type: ignore
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
